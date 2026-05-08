"""
Entry point for function-level repair.

Constructs a RepairInput from command-line arguments or defaults,
then calls the repair() function.

Usage:
    python divergence_guided_repair.py --project_path ../tests/fmt \\
                   --project_name fmt \\
                   --test_path ../tests/fmt/test/chrono-test.cc \\
                   --test_case_name "YearMonthDay" \\
                   --test_start_line 1020 \\
                   --test_end_line 1052 \\
                   --root_cause_file ../tests/fmt/include/fmt/chrono.h \\
                   --root_cause_name "strftime" \\
                   --root_cause_start 350 \\
                   --root_cause_end 420
"""

import argparse
import json
import os
import shutil
import sys
from datetime import datetime

from repair.Models import (
    FunctionInfo, TypeMetadata, TypeDepInfo,
    TestInfo, RepairInput, RepairResult,
)
from repair.Repairer import repair
from repair.WorkflowConfig import WorkflowConfig, default_config, load_config


def load_function_info_from_source(file_path: str, start_line: int, end_line: int, name: str) -> FunctionInfo:
    """Load a FunctionInfo by reading source lines from a file."""
    with open(file_path, 'r') as f:
        all_lines = f.readlines()
    source = "".join(all_lines[start_line - 1:end_line])
    return FunctionInfo(
        name=name,
        file_path=os.path.abspath(file_path),
        start_line=start_line,
        end_line=end_line,
        source_code=source,
    )


def load_repair_input_from_json(config_path: str) -> RepairInput:
    """Load a RepairInput from a JSON configuration file.

    Expected JSON format:
    {
        "project_name": "fmt",
        "project_path": "../tests/fmt",
        "test_info": {
            "test_path": "../tests/fmt/test/chrono-test.cc",
            "test_case_name": "YearMonthDay",
            "test_start_line": 1020,
            "test_end_line": 1052,
            "initial_failure_output": "..."
        },
        "root_cause_function": {
            "name": "strftime",
            "file_path": "../tests/fmt/include/fmt/chrono.h",
            "start_line": 350,
            "end_line": 420
        },
        "candidate_functions": [
            {
                "name": "format_tm",
                "file_path": "../tests/fmt/include/fmt/chrono.h",
                "start_line": 200,
                "end_line": 250
            }
        ],
        "type_deps": {
            "skeleton_graph": {"tm_wrapper": ["tm"]},
            "full_definitions": {"tm_wrapper": "struct tm_wrapper { ... }"},
            "type_metadata": {
                "tm_wrapper": {
                    "qualified_name": "fmt::tm_wrapper",
                    "file_path": "...",
                    "start_line": 100,
                    "end_line": 120
                }
            }
        },
        "max_iterations": 30,
        "model": "gemini-2.5-flash"
    }
    """
    with open(config_path, 'r') as f:
        config = json.load(f)

    # Build TestInfo
    ti = config["test_info"]
    test_info = TestInfo(
        test_path=os.path.abspath(ti["test_path"]),
        test_case_name=ti["test_case_name"],
        test_start_line=ti["test_start_line"],
        test_end_line=ti["test_end_line"],
        initial_failure_output=ti.get("initial_failure_output", ""),
        failing_statement=ti.get("failing_statement", ""),
    )

    # Build root cause FunctionInfo
    rc = config["root_cause_function"]
    root_cause = load_function_info_from_source(
        rc["file_path"], rc["start_line"], rc["end_line"], rc["name"]
    )

    # Build candidate FunctionInfos
    candidates = []
    for c in config.get("candidate_functions", []):
        candidates.append(load_function_info_from_source(
            c["file_path"], c["start_line"], c["end_line"], c["name"]
        ))

    # Build TypeDepInfo
    td = config.get("type_deps", {})
    skeleton_graph = td.get("skeleton_graph", {})
    full_definitions = td.get("full_definitions", {})

    type_metadata = {}
    for name, meta in td.get("type_metadata", {}).items():
        type_metadata[name] = TypeMetadata(
            qualified_name=meta["qualified_name"],
            file_path=meta.get("file_path", ""),
            start_line=meta.get("start_line", 0),
            end_line=meta.get("end_line", 0),
        )

    type_deps = TypeDepInfo(
        skeleton_graph=skeleton_graph,
        full_definitions=full_definitions,
        type_metadata=type_metadata,
    )

    return RepairInput(
        project_name=config["project_name"],
        project_path=os.path.abspath(config["project_path"]),
        test_info=test_info,
        root_cause_function=root_cause,
        caller_functions=[],
        suspicious_functions=candidates,
        type_deps=type_deps,
        max_iterations=config.get("max_iterations", 30),
        model=config.get("model", "gemini-2.5-flash"),
        compile_script=config.get("compile_script", "compile.sh"),
        run_script=config.get("run_script", "run.sh"),
    )


# ---------------------------------------------------------------------------
# trace_analysis.json loader
# ---------------------------------------------------------------------------

def _extract_short_name(full_name: str) -> str:
    """Extract short function name from a full signature.

    'system_strftime(const std::string &, ...)' -> 'system_strftime'
    'tao::pegtl::internal::mmap_file_posix' -> 'mmap_file_posix'
    'mmap_file_posix::<constructor>' -> 'mmap_file_posix'
    """
    paren = full_name.find('(')
    if paren >= 0:
        full_name = full_name[:paren]
    parts = full_name.split('::')
    name = parts[-1].strip()
    # For constructors/destructors, use the class name instead
    if name in ('<constructor>', '<destructor>'):
        return parts[-2].strip() if len(parts) >= 2 else name
    return name


def _find_function_end_line(file_path: str, start_line: int) -> int:
    """Estimate the end line of a C/C++ function by brace counting.

    Returns 0 if the file cannot be read or no matching braces are found,
    so callers can tell the estimate failed rather than getting a made-up value.
    """
    try:
        with open(file_path, 'r') as f:
            lines = f.readlines()
    except OSError:
        return 0

    depth = 0
    found_open = False
    for i in range(start_line - 1, min(len(lines), start_line + 500)):
        for ch in lines[i]:
            if ch == '{':
                depth += 1
                found_open = True
            elif ch == '}':
                depth -= 1
                if found_open and depth == 0:
                    return i + 1  # 1-based
    return 0


def _build_failure_description(trace_data: dict) -> str:
    """Synthesize initial_failure_output from trace analysis comparison data.

    Only includes fields that are actually present in the data.
    Returns empty string when nothing is available.
    """
    analysis_type = trace_data.get("analysis_type", "")
    parts = []

    if analysis_type == "function_boundary_analysis":
        rc = trace_data.get("root_cause", {})
        comp = rc.get("comparison", {})
        if rc.get("function_name"):
            parts.append(f"Root cause: {rc['function_name']}")
        if rc.get("detection_method"):
            parts.append(f"Detection: {rc['detection_method']}")
        if rc.get("discrepancy_reason"):
            parts.append(f"Discrepancy reason from preprocessing: {rc['discrepancy_reason']}")
        if comp:
            # Show inputs only when both sides have non-empty values.
            # Empty inputs ({}) are a preprocessing artifact — the actual
            # inputs are assumed identical for wasm and native.

            parts.append(f"Inputs consistent: {comp.get('inputs_consistent', True)}")
            native_inputs = [i for i in comp.get("native_inputs", []) if i]
            wasm_inputs = [i for i in comp.get("wasm_inputs", []) if i]
            if native_inputs and wasm_inputs:
                for inp in native_inputs:
                    parts.append(f"Native input: {inp}")
                for inp in wasm_inputs:
                    parts.append(f"Wasm input: {inp}")
            if "outputs_consistent" in comp:
                parts.append(f"Outputs consistent: {comp['outputs_consistent']}")
            for out in comp.get("native_outputs", []):
                parts.append(f"Native output: {out}")
            for out in comp.get("wasm_outputs", []):
                parts.append(f"Wasm output: {out}")

    elif analysis_type == "dynamic_trace_analysis":
        rcs = trace_data.get("root_causes", [])
        if rcs:
            rc = rcs[0]
            if rc.get("function_name"):
                parts.append(f"Root cause: {rc['function_name']}")
            if rc.get("discrepancy_type"):
                parts.append(f"Discrepancy type: {rc['discrepancy_type']}")
            if rc.get("native_summary"):
                parts.append(f"Native: {rc['native_summary']}")
            if rc.get("wasm_summary"):
                parts.append(f"Wasm: {rc['wasm_summary']}")

    return "\n".join(parts)


def _load_type_deps_from_function_types_data(data: dict) -> TypeDepInfo:
    """Parse type dependencies from function_types data (dict).

    Flattens the per-file per-function type mapping into the flat
    skeleton_graph / full_definitions / type_metadata structure.
    """
    skeleton_graph = {}
    full_definitions = {}
    type_metadata_dict = {}

    for _file_path, functions in data.items():
        if not isinstance(functions, dict):
            continue
        for _func_name, func_data in functions.items():
            if not isinstance(func_data, dict):
                continue
            for type_file, types in func_data.get("types", {}).items():
                if not isinstance(types, dict):
                    continue
                for type_qname, type_info in types.items():
                    if not isinstance(type_info, dict):
                        continue
                    if type_qname in type_metadata_dict:
                        continue  # already collected
                    skeleton_graph[type_qname] = type_info.get("depends_on", [])
                    full_definitions[type_qname] = type_info.get("body", "")
                    loc = type_info.get("location", {})
                    type_metadata_dict[type_qname] = TypeMetadata(
                        qualified_name=type_qname,
                        file_path=type_file,
                        start_line=loc.get("start_line", 0),
                        end_line=loc.get("end_line", 0),
                    )

    return TypeDepInfo(
        skeleton_graph=skeleton_graph,
        full_definitions=full_definitions,
        type_metadata=type_metadata_dict,
    )


def _load_type_deps_from_function_types(function_types_path: str) -> TypeDepInfo:
    """Load TypeDepInfo from a function_types.json file (fallback)."""
    with open(function_types_path, 'r') as f:
        data = json.load(f)
    return _load_type_deps_from_function_types_data(data)


def load_repair_input_from_trace_analysis(
    trace_analysis_path: str,
    project_path: str = None,
    test_case_name: str = "",
    max_iterations: int = 30,
    max_tokens: int = 65536,
    model: str = "gemini-2.5-flash",
) -> RepairInput:
    """Load a RepairInput from a preprocessing trace_analysis.json.

    Handles both analysis types produced by preprocessing:
      - function_boundary_analysis (e.g. fmt): root_cause + suspicious_func_list
      - dynamic_trace_analysis (e.g. PEGTL): root_causes + symptoms

    Auto-discovers sibling files in the same directory:
      - function_types.json  — type dependencies for involved functions

    Args:
        trace_analysis_path:  Path to trace_analysis.json.
        project_path:         Project root (default: directory of trace_analysis.json).
        test_case_name:       Failing test case name for log filtering.
        max_iterations:       Max repair iterations.
        model:                LLM model identifier.
    """
    trace_analysis_path = os.path.abspath(trace_analysis_path)
    trace_dir = os.path.dirname(trace_analysis_path)

    with open(trace_analysis_path, 'r') as f:
        trace_data = json.load(f)

    # Derive project_path
    if not project_path:
        project_path = trace_dir
    project_path = os.path.abspath(project_path)
    project_name = os.path.basename(project_path)

    analysis_type = trace_data.get("analysis_type", "")

    # --- Test info ---
    # Test case info from trace_analysis.json
    tfi = trace_data.get("test_case", {})

    test_path = tfi.get("file_path", "")
    # Key names vary across projects / format versions
    test_start_line = int(tfi.get("test case start line",
                          tfi.get("line",
                          tfi.get("failed line", "0"))))
    test_end_line = int(tfi.get("test case end line", "0"))

    if test_path and test_start_line > 0 and test_end_line == 0:
        test_end_line = _find_function_end_line(test_path, test_start_line)

    initial_failure_output = _build_failure_description(trace_data)

    test_info = TestInfo(
        test_path=os.path.abspath(test_path) if test_path else "",
        test_case_name=test_case_name,
        test_start_line=test_start_line,
        test_end_line=test_end_line,
        initial_failure_output=initial_failure_output,
    )

    # --- Root cause function ---
    if analysis_type == "function_boundary_analysis":
        rc = trace_data["root_cause"]
    else:
        rcs = trace_data.get("root_causes", [])
        if not rcs:
            raise ValueError("No root causes found in trace_analysis.json")
        rc = rcs[0]

    # Preprocessing may not identify a specific root-cause function — e.g.,
    # test failures caused by environment/setup issues where the bug is not
    # in a called function. In that case, fall back to treating the test
    # case itself as the "root cause", so the LLM sees the test source and
    # is guided (by the existing "do NOT modify tests" rule) to look for
    # fixes in build config, environment setup, or library code.
    rc_file_raw = rc.get("file") or rc.get("file_path") or ""
    rc_line_raw = rc.get("line") or 0
    rc_name_raw = (rc.get("function_name") or "").strip()

    try:
        rc_line_int = int(rc_line_raw) if rc_line_raw else 0
    except (TypeError, ValueError):
        rc_line_int = 0

    has_specific_rc = bool(
        rc_file_raw and rc_line_int and rc_name_raw and os.path.isfile(rc_file_raw)
    )

    if has_specific_rc:
        rc_file = rc_file_raw
        rc_start = rc_line_int
        rc_end = rc.get("end_line") or _find_function_end_line(rc_file, rc_start) or rc_start
        rc_name = _extract_short_name(rc_name_raw)
    else:
        # Fall back to test case as root cause.
        print("No specific root-cause function in trace analysis — "
              "using the failing test case as the root cause.")
        if not (test_path and test_start_line > 0):
            raise ValueError(
                "Cannot derive root cause: trace analysis has no root cause "
                "function AND no valid test case location."
            )
        rc_file = os.path.abspath(test_path)
        rc_start = test_start_line
        rc_end = test_end_line if test_end_line > 0 else (
            _find_function_end_line(rc_file, rc_start) or rc_start
        )
        rc_name = os.path.splitext(os.path.basename(rc_file))[0]

    root_cause = load_function_info_from_source(rc_file, rc_start, rc_end, rc_name)

    # --- Caller functions (execution chain from remaining_stack) ---
    caller_functions = []
    seen = set()
    remaining_stack = trace_data.get("remaining_stack", [])
    for entry in remaining_stack:
        sf_file = entry.get("file_path", "")
        sf_line = entry.get("line", 0)
        if not sf_file or sf_line == 0:
            continue
        key = (sf_file, sf_line)
        if key in seen or (sf_file == rc_file and sf_line == rc_start):
            continue
        seen.add(key)
        sf_end = entry.get("end_line") or _find_function_end_line(sf_file, sf_line) or sf_line
        sf_name = _extract_short_name(entry.get("function_name", entry.get("func_name", "unknown")))
        try:
            caller_functions.append(
                load_function_info_from_source(sf_file, sf_line, sf_end, sf_name)
            )
        except (OSError, IndexError):
            continue

    # --- Suspicious functions (potentially related, role uncertain) ---
    suspicious_functions = []
    if analysis_type == "function_boundary_analysis":
        raw_suspicious = trace_data.get("suspicious_func_list", [])
    else:
        raw_suspicious = trace_data.get("symptoms", [])

    for entry in raw_suspicious:
        sf_file = entry.get("file", "")
        sf_line = entry.get("line", 0)
        if not sf_file or sf_line == 0:
            continue
        key = (sf_file, sf_line)
        if key in seen or (sf_file == rc_file and sf_line == rc_start):
            continue
        seen.add(key)
        sf_end = entry.get("end_line") or _find_function_end_line(sf_file, sf_line) or sf_line
        sf_name = _extract_short_name(entry.get("function_name", entry.get("func_name", "unknown")))
        try:
            suspicious_functions.append(
                load_function_info_from_source(sf_file, sf_line, sf_end, sf_name)
            )
        except (OSError, IndexError):
            continue

    # --- Type dependencies ---
    # Prefer type_info embedded in trace_analysis.json (per test case, always in sync).
    # Fall back to separate function_types.json file if not provided.
    type_info_data = trace_data.get("type_info", {})
    if type_info_data:
        type_deps = _load_type_deps_from_function_types_data(type_info_data)
    else:
        print("Loading type information from function_types.json (not embedded in trace_analysis.json)")
        function_types_path = os.path.join(trace_dir, "function_types.json")
        if os.path.exists(function_types_path):
            type_deps = _load_type_deps_from_function_types(function_types_path)
        else:
            type_deps = TypeDepInfo(
                skeleton_graph={}, full_definitions={}, type_metadata={},
            )

    print(f"Loaded trace analysis ({analysis_type}) from {trace_analysis_path}")
    print(f"  Project: {project_name} ({project_path})")
    print(f"  Root cause: {rc_name}() at {root_cause.location_str()}")
    print(f"  Caller functions: {len(caller_functions)}")
    print(f"  Suspicious functions: {len(suspicious_functions)}")
    print(f"  Type deps: {len(type_deps.type_metadata)} types")
    print(f"  Test: {test_info.test_path}:{test_info.test_start_line}-{test_info.test_end_line}")

    return RepairInput(
        project_name=project_name,
        project_path=project_path,
        test_info=test_info,
        root_cause_function=root_cause,
        caller_functions=caller_functions,
        suspicious_functions=suspicious_functions,
        type_deps=type_deps,
        max_iterations=max_iterations,
        max_tokens=max_tokens,
        model=model,
    )


def print_result(result: RepairResult):
    """Print a human-readable summary of the repair result."""
    print("\n" + "=" * 60)
    print(f"REPAIR RESULT: {result.status.upper()}")
    print(f"Final state: {result.final_state}")
    print(f"Iterations used: {result.iterations_used}")
    print(f"Total tokens: {result.total_tokens} (repair: {result.repair_tokens} [in:{result.repair_input_tokens}/out:{result.repair_output_tokens}], instrumentation: {result.instrumentation_tokens} [in:{result.instrumentation_input_tokens}/out:{result.instrumentation_output_tokens}])")
    print(f"Reason: {result.reason}")

    if result.fix:
        print(f"\nSuccessful fix:")
        print(f"  File: {result.fix.file_path}")
        print(f"  Description: {result.fix.description}")
        print(f"  Original code:\n{_indent(result.fix.original_code, 4)}")
        print(f"  Fixed code:\n{_indent(result.fix.fixed_code, 4)}")

    print(f"\nPatch attempts: {len(result.history.fix_attempts)}")
    print(f"Tool calls: {len(result.history.tool_calls)}")
    print(f"Instrumentation runs: {len(result.history.instrumentation_results)}")
    print("=" * 60)


def _indent(text: str, spaces: int) -> str:
    prefix = " " * spaces
    return "\n".join(prefix + line for line in text.splitlines())


def main():
    parser = argparse.ArgumentParser(
        description="Function-level Wasm discrepancy repair."
    )
    parser.add_argument(
        "--config", type=str, default=None,
        help="Path to a JSON config file (overrides all other arguments)."
    )
    parser.add_argument(
        "--trace_analysis", type=str, default=None,
        help="Path to trace_analysis.json from preprocessing. "
             "Auto-discovers metadata.json and function_types.json in the same "
             "directory. Use --project_path, --test_case_name "
             "to override auto-discovered values."
    )
    parser.add_argument("--project_path", type=str, default=None)
    parser.add_argument("--project_name", type=str, default=None)
    parser.add_argument("--test_path", type=str, default=None)
    parser.add_argument("--test_case_name", type=str, default="")
    parser.add_argument("--test_start_line", type=int, default=0)
    parser.add_argument("--test_end_line", type=int, default=0)
    parser.add_argument("--initial_failure_output", type=str, default="")
    parser.add_argument("--failing_statement", type=str, default="")
    parser.add_argument("--root_cause_file", type=str, default=None)
    parser.add_argument("--root_cause_name", type=str, default="unknown")
    parser.add_argument("--root_cause_start", type=int, default=0)
    parser.add_argument("--root_cause_end", type=int, default=0)
    parser.add_argument("--max_iterations", type=int, default=50)
    parser.add_argument("--max_tokens", type=int, default=8192)
    parser.add_argument("--model", type=str, default="gemini-3-flash-preview")
    parser.add_argument(
        "--workflow_config", type=str, default=None,
        help="Path to a workflow config JSON file (controls available actions and "
             "provided information for ablation studies). If not specified, uses "
             "the default config with all tools and full information."
    )
    # Ablation flags (override workflow_config flags if both specified)
    parser.add_argument(
        "--no_trace_analysis", action="store_true", default=False,
        help="Ablation: hide all trace analysis info (root cause function, "
             "discrepancy description). Only test location is shown.")
    parser.add_argument(
        "--no_candidates", action="store_true", default=False,
        help="Ablation: hide candidate functions.")
    parser.add_argument(
        "--no_instrumentation", action="store_true", default=False,
        help="Ablation: remove instrumentation tools.")
    parser.add_argument(
        "--no_type_deps", action="store_true", default=False,
        help="Ablation: remove type dependency tools.")
    # Experiment management
    parser.add_argument(
        "--restore", action="store_true", default=False,
        help="Before repair, auto-restore source files from previous run's backup."
    )
    parser.add_argument(
        "--collect", action="store_true", default=False,
        help="After repair, append results to results/{project}/{test_case}/results.xlsx."
    )
    parser.add_argument(
        "--results_dir", type=str, default="results",
        help="Base directory for result output (default: results)."
    )
    parser.add_argument(
        "--always_restore", action="store_true", default=False,
        help="After repair completes, restore all modified source files to "
             "their original state (regardless of patch success). Ensures a "
             "clean source tree between experiment runs."
    )
    parser.add_argument(
        "--clean_build_dir", action="store_true", default=False,
        help="Before pre-repair compile, remove build_wasm/ and build_native/ "
             "entirely (not just test/ subdirs). Forces a full clean rebuild. "
             "Slower but avoids stale CMake cache or library object files."
    )
    args = parser.parse_args()

    # Restore files from previous run BEFORE loading repair input,
    # so that source files on disk are clean when we read them.
    if args.restore:
        # Derive project name and test case from args to find the results directory
        if args.trace_analysis:
            project_name = os.path.basename(
                os.path.abspath(args.project_path or os.path.dirname(args.trace_analysis)))
        elif args.project_name:
            project_name = args.project_name
        elif args.project_path:
            project_name = os.path.basename(os.path.abspath(args.project_path))
        else:
            project_name = "unknown"
        test_case = args.test_case_name or "default"
        project_results_dir = os.path.join(args.results_dir, project_name, test_case)
        _auto_restore(project_results_dir)

    # Load repair input
    if args.trace_analysis:
        repair_input = load_repair_input_from_trace_analysis(
            args.trace_analysis,
            project_path=args.project_path,
            test_case_name=args.test_case_name,
            max_iterations=args.max_iterations,
            max_tokens=args.max_tokens,
            model=args.model,
        )
    elif args.config:
        repair_input = load_repair_input_from_json(args.config)
    else:
        # Build from CLI args
        if not args.project_path or not args.root_cause_file:
            parser.error("Either --config or --project_path + --root_cause_file is required.")

        root_cause = load_function_info_from_source(
            args.root_cause_file, args.root_cause_start, args.root_cause_end, args.root_cause_name
        )

        test_info = TestInfo(
            test_path=os.path.abspath(args.test_path) if args.test_path else "",
            test_case_name=args.test_case_name,
            test_start_line=args.test_start_line,
            test_end_line=args.test_end_line,
            initial_failure_output=args.initial_failure_output,
            failing_statement=args.failing_statement,
        )

        repair_input = RepairInput(
            project_name=args.project_name or os.path.basename(args.project_path),
            project_path=os.path.abspath(args.project_path),
            test_info=test_info,
            root_cause_function=root_cause,
            caller_functions=[],
            suspicious_functions=[],
            type_deps=TypeDepInfo(
                skeleton_graph={},
                full_definitions={},
                type_metadata={},
            ),
            max_iterations=args.max_iterations,
            max_tokens=args.max_tokens,
            model=args.model,
        )

    # Load workflow config
    if args.workflow_config:
        workflow_config = load_config(args.workflow_config)
    else:
        workflow_config = default_config()

    # Apply CLI ablation flags (override config)
    if args.no_trace_analysis:
        workflow_config.provide_trace_analysis = False
    if args.no_candidates:
        workflow_config.provide_candidates = False
    if args.no_instrumentation:
        workflow_config.provide_instrumentation = False
    if args.no_type_deps:
        workflow_config.provide_type_deps = False
    # Re-run __post_init__ to apply flag-driven action adjustments
    workflow_config.__post_init__()

    # Build ablation tag for output directory
    tag_parts = []
    if not workflow_config.provide_trace_analysis:
        tag_parts.append("no_trace")
    if not workflow_config.provide_candidates:
        tag_parts.append("no_candidates")
    if not workflow_config.provide_instrumentation:
        tag_parts.append("no_instrument")
    if not workflow_config.provide_type_deps:
        tag_parts.append("no_type_deps")
    ablation_tag = "_".join(tag_parts) if tag_parts else "default"

    # Derive project results dir (restore already happened above if --restore)
    test_case_name = repair_input.test_info.test_case_name or "default"
    project_results_dir = os.path.join(args.results_dir, repair_input.project_name, test_case_name)

    # Compile the project to ensure binaries match current source
    _pre_repair_compile(repair_input, clean_build_dir=args.clean_build_dir)

    # Create output directory: results/{project_name}/{test_case}/run_{YYYYMMDD_HHMMSS}_{model}_{tag}/
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    model_tag = repair_input.model.replace("/", "_")
    output_dir = os.path.join(project_results_dir,
                              f"run_{timestamp}_{model_tag}_{ablation_tag}")
    os.makedirs(output_dir, exist_ok=True)

    print(f"Output directory: {output_dir}")
    print(f"Ablation config: {ablation_tag}")
    print(f"  provide_trace_analysis={workflow_config.provide_trace_analysis}")
    print(f"  provide_candidates={workflow_config.provide_candidates}")
    print(f"  provide_instrumentation={workflow_config.provide_instrumentation}")
    print(f"  provide_type_deps={workflow_config.provide_type_deps}")
    print(f"  ANALYZE actions: {sorted(workflow_config.analyze_actions.keys())}")
    print(f"  PATCH actions: {sorted(workflow_config.patch_actions.keys())}")

    # Save config snapshot for reproducibility
    config_snapshot = {
        "provide_trace_analysis": workflow_config.provide_trace_analysis,
        "provide_candidates": workflow_config.provide_candidates,
        "provide_instrumentation": workflow_config.provide_instrumentation,
        "provide_type_deps": workflow_config.provide_type_deps,
        "analyze_actions": sorted(workflow_config.analyze_actions.keys()),
        "patch_actions": sorted(workflow_config.patch_actions.keys()),
        "model": repair_input.model,
        "max_iterations": repair_input.max_iterations,
    }
    with open(os.path.join(output_dir, "config.json"), 'w') as f:
        json.dump(config_snapshot, f, indent=2)

    # Run repair
    result = repair(repair_input, workflow_config=workflow_config,
                    output_dir=output_dir)
    print_result(result)

    # Save result as JSON
    result_path = os.path.join(output_dir, "repair_result.json")
    _save_result_json(result, result_path)
    print(f"\nResult saved to {result_path}")

    # Append to project-level Excel spreadsheet if --collect is set
    if args.collect:
        excel_path = os.path.join(project_results_dir, "results.xlsx")
        run_name = os.path.basename(output_dir)
        _append_to_excel(excel_path, run_name, ablation_tag, repair_input.model, test_case_name, result)

    # If --always_restore, revert any source modifications (even on success).
    # Reuses _auto_restore which reads {project_results_dir}/latest_restore/
    # (populated during this run by Toolkit._backup_file with ORIGINAL content).
    if args.always_restore:
        print("\n--always_restore: reverting any source modifications.")
        _auto_restore(project_results_dir)


def _append_to_excel(excel_path: str, run_name: str, config_tag: str,
                     model: str, test_case_name: str, result: RepairResult):
    """Append one row of results to the project-level Excel spreadsheet.

    Each test case gets its own sheet. Creates the file and/or sheet
    with headers if they don't exist.
    """
    try:
        from openpyxl import Workbook, load_workbook
    except ImportError:
        print("openpyxl not installed — skipping Excel output. "
              "Install with: pip install openpyxl")
        return

    headers = [
        "run_dir", "config_tag", "model", "status", "iterations_used",
        "total_tokens", "repair_tokens", "repair_input_tokens", "repair_output_tokens",
        "instrumentation_tokens", "instrumentation_input_tokens", "instrumentation_output_tokens",
        "fix_attempts_count", "tool_calls_count", "instrumentation_count",
    ]

    row = [
        run_name,
        config_tag,
        model,
        result.status,
        result.iterations_used,
        result.total_tokens,
        result.repair_tokens,
        result.repair_input_tokens,
        result.repair_output_tokens,
        result.instrumentation_tokens,
        result.instrumentation_input_tokens,
        result.instrumentation_output_tokens,
        len(result.history.fix_attempts),
        len(result.history.tool_calls),
        len(result.history.instrumentation_results),
    ]

    # Sanitize sheet name (Excel limits: 31 chars, forbids \ / * ? : [ ])
    import re
    sheet_name = re.sub(r'[\\/*?\[\]:]', '_', test_case_name)[:31]

    if os.path.exists(excel_path):
        wb = load_workbook(excel_path)
    else:
        wb = Workbook()
        # Remove the default empty sheet
        wb.remove(wb.active)

    # Get or create the sheet for this test case
    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        ws = wb.create_sheet(title=sheet_name)
        ws.append(headers)

    ws.append(row)
    wb.save(excel_path)
    print(f"Results appended to {excel_path} (sheet: {sheet_name})")


def _pre_repair_compile(repair_input: RepairInput, clean_build_dir: bool = False):
    """Compile the project before repair to ensure binaries match source.

    When clean_build_dir is True, removes build_wasm/ and build_native/ entirely
    (full clean rebuild). Otherwise, only removes the test/ subdirs within them
    (partial clean — faster, relies on CMake incremental build).
    """
    import subprocess

    script = repair_input.compile_script
    compile_script = (script if os.path.isabs(script)
                      else os.path.join(repair_input.project_path, script))

    if not os.path.exists(compile_script):
        print(f"WARNING: compile script not found: {compile_script} — skipping pre-compile")
        return

    print(f"Pre-repair compile: {compile_script}")
    try:
        # Clean build artifacts before compiling. Two modes:
        #   - clean_build_dir=True:  remove entire build_wasm/ and build_native/
        #   - clean_build_dir=False: only remove test/ subdirs (existing default)
        if clean_build_dir:
            dirs_to_clean = [
                os.path.join(repair_input.project_path, "build_wasm"),
                os.path.join(repair_input.project_path, "build_native"),
            ]
            label = "ENTIRE build dir (--clean_build_dir)"
        else:
            dirs_to_clean = [
                os.path.join(repair_input.project_path, "build_wasm", "test"),
                os.path.join(repair_input.project_path, "build_native", "test"),
            ]
            label = "stale test binaries"
        for d in dirs_to_clean:
            if os.path.exists(d):
                shutil.rmtree(d)
                print(f"  Cleaned up {d} ({label})")
        result = subprocess.run(
            ["bash", os.path.abspath(compile_script)],
            cwd=repair_input.project_path,
            stdout=subprocess.PIPE,
            stderr=subprocess.PIPE,
            text=True,
            timeout=300,
        )
        if result.returncode == 0:
            print("  Compile: OK")
        else:
            print(f"  Compile: FAILED (exit code {result.returncode})")
            print(f"  {result.stderr[-200:]}" if result.stderr else "")
    except subprocess.TimeoutExpired:
        print("  Compile: TIMEOUT (300s)")
        exit(1)
    except Exception as e:
        print(f"  Compile: ERROR ({e})")
        exit(1)
    print()


def _auto_restore(project_results_dir: str):
    """Auto-restore source files from a previous run's latest_restore/ backup.

    If results/{proj}/latest_restore/ exists with a manifest, restore all
    files listed in it, then delete the latest_restore/ directory.
    """
    latest_dir = os.path.join(project_results_dir, "latest_restore")
    manifest_path = os.path.join(latest_dir, "manifest.json")

    if not os.path.exists(manifest_path):
        return

    with open(manifest_path, 'r') as f:
        manifest = json.load(f)

    if not manifest:
        return

    print("Auto-restoring files from previous run...")
    for safe_name, original_path in manifest.items():
        backup_path = os.path.join(latest_dir, safe_name)
        if not os.path.exists(backup_path):
            print(f"  WARNING: backup missing: {backup_path}")
            continue
        with open(backup_path, 'r') as f:
            content = f.read()
        with open(original_path, 'w') as f:
            f.write(content)
        print(f"  Restored: {original_path}")

    # Clean up latest_restore/ so it doesn't trigger again
    shutil.rmtree(latest_dir)
    print(f"  Cleaned up {latest_dir}")
    print()

    # Recompile the repo
    


def _save_result_json(result: RepairResult, path: str):
    """Save the RepairResult to a JSON file."""
    import dataclasses

    def _serialize(obj):
        if dataclasses.is_dataclass(obj) and not isinstance(obj, type):
            return dataclasses.asdict(obj)
        if hasattr(obj, '__dict__'):
            return obj.__dict__
        return str(obj)

    data = {
        "status": result.status,
        "reason": result.reason,
        "final_state": result.final_state,
        "iterations_used": result.iterations_used,
        "total_tokens": result.total_tokens,
        "repair_tokens": result.repair_tokens,
        "repair_input_tokens": result.repair_input_tokens,
        "repair_output_tokens": result.repair_output_tokens,
        "instrumentation_tokens": result.instrumentation_tokens,
        "instrumentation_input_tokens": result.instrumentation_input_tokens,
        "instrumentation_output_tokens": result.instrumentation_output_tokens,
        "fix_attempts_count": len(result.history.fix_attempts),
        "tool_calls_count": len(result.history.tool_calls),
        "instrumentation_count": len(result.history.instrumentation_results),
    }

    if result.fix:
        data["fix"] = {
            "file_path": result.fix.file_path,
            "description": result.fix.description,
            "original_code": result.fix.original_code,
            "fixed_code": result.fix.fixed_code,
        }

    # Full history for reproducibility
    data["history"] = {
        "current_plan": result.history.current_plan,
        "tool_calls": [
            {"tool_name": tc.tool_name, "params": tc.params,
             "result_summary": tc.result_summary}
            for tc in result.history.tool_calls
        ],
        "fix_attempts": [
            {"description": fa.description, "file_path": fa.file_path,
             "original_code": fa.original_code, "fixed_code": fa.fixed_code,
             "compile_success": fa.compile_success, "test_passed": fa.test_passed,
             "error_output": fa.error_output,
             "bypass_rejected": fa.bypass_rejected,
             "rejection_reason": fa.rejection_reason,
             "analysis": fa.analysis}
            for fa in result.history.fix_attempts
        ],
        "instrumentation_results": result.history.instrumentation_results,
    }

    with open(path, 'w') as f:
        json.dump(data, f, indent=2, default=_serialize)


if __name__ == "__main__":
    main()
